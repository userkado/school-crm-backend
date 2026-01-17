from typing import Annotated
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from openpyxl import load_workbook
import io

from app.db.session import get_db
from app.models.school import Student, ClassGroup
from app.schemas.school import StudentCreate, StudentResponse
from app.api.deps import allow_admin, get_current_user

router = APIRouter()

# --- 1. СОЗДАТЬ ОДНОГО (Обычный метод) ---
@router.post("/", response_model=StudentResponse)
async def create_student(
    student_in: StudentCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    _ = Depends(allow_admin)
):
    new_student = Student(
        full_name=student_in.full_name,
        class_group_id=student_in.class_group_id
    )
    db.add(new_student)
    await db.commit()
    await db.refresh(new_student)
    return new_student

# --- 2. МАССОВАЯ ЗАГРУЗКА (EXCEL) ---
@router.post("/upload")
async def upload_students_excel(
    class_id: int, # В какой класс добавляем всех из файла
    file: UploadFile = File(...),
    db: Annotated[AsyncSession, Depends(get_db)] = None, # = None заглушка
    _ = Depends(allow_admin)
):
    """
    Читает Excel файл. Ожидается, что в Колонке A - Имена учеников.
    Все ученики из файла будут добавлены в указанный class_id.
    """
    # Проверяем, существует ли класс
    class_group = await db.get(ClassGroup, class_id)
    if not class_group:
        raise HTTPException(status_code=404, detail="Class not found")

    # Читаем файл
    contents = await file.read()
    wb = load_workbook(filename=io.BytesIO(contents))
    sheet = wb.active
    
    count = 0
    # Читаем строки (пропускаем заголовок, если он есть - начнем со 2 строки? 
    # Нет, давайте читать всё с 1-й, если там просто список имен)
    for row in sheet.iter_rows(values_only=True):
        full_name = row[0] # Колонка А
        if full_name:
            new_student = Student(full_name=str(full_name), class_group_id=class_id)
            db.add(new_student)
            count += 1
    
    await db.commit()
    return {"message": f"Успешно добавлено {count} учеников в класс {class_group.name}"}

# --- 3. ПОЛУЧИТЬ СПИСОК (С фильтром по классу) ---
@router.get("/", response_model=list[StudentResponse])
async def get_students(
    db: Annotated[AsyncSession, Depends(get_db)],
    _ = Depends(get_current_user),
    class_id: int | None = None
):
    query = select(Student)
    if class_id:
        query = query.filter(Student.class_group_id == class_id)
    result = await db.execute(query)
    return result.scalars().all()

# --- 4. ПЕРЕВОД УЧЕНИКА (Трансфер) ---
@router.put("/{student_id}/transfer")
async def transfer_student(
    student_id: int,
    new_class_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    _ = Depends(allow_admin)
):
    # Ищем ученика
    student = await db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Ищем новый класс
    new_class = await db.get(ClassGroup, new_class_id)
    if not new_class:
        raise HTTPException(status_code=404, detail="Target Class not found")
        
    # Переводим
    student.class_group_id = new_class_id
    await db.commit()
    return {"message": f"Ученик переведен в {new_class.name}"}
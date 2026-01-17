from typing import Annotated
from datetime import date
import io
from urllib.parse import quote  # 👈 1. ДОБАВЛЕН ВАЖНЫЙ ИМПОРТ

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.session import get_db
from app.models.school import Student, Grade, Attendance, ClassGroup
from app.api.deps import allow_teacher

router = APIRouter()

# --- ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ ---
async def get_report_data(class_id: int, start_date: date, end_date: date, report_type: str, db: AsyncSession):
    res_st = await db.execute(select(Student).filter(Student.class_group_id == class_id).order_by(Student.full_name))
    students = res_st.scalars().all()
    
    data = []

    for s in students:
        row = {"full_name": s.full_name}
        
        if report_type == "grades":
            query = select(Grade).filter(
                Grade.student_id == s.id,
                Grade.date >= start_date,
                Grade.date <= end_date
            )
            res_grades = await db.execute(query)
            grades = res_grades.scalars().all()
            
            if grades:
                avg = sum(g.value for g in grades) / len(grades)
                row["value"] = round(avg, 2)
                row["count"] = len(grades)
            else:
                row["value"] = 0
                row["count"] = 0

        elif report_type == "attendance":
            query = select(Attendance).filter(
                Attendance.student_id == s.id,
                Attendance.date >= start_date,
                Attendance.date <= end_date
            )
            res_att = await db.execute(query)
            atts = res_att.scalars().all()
            
            row["absent"] = sum(1 for a in atts if a.status == 'ABSENT')
            row["late"] = sum(1 for a in atts if a.status == 'LATE')
        
        data.append(row)
    
    return data

# --- 1. JSON ОТЧЕТ ---
@router.get("/view")
async def view_report(
    class_id: int,
    report_type: str,
    start_date: date,
    end_date: date,
    db: Annotated[AsyncSession, Depends(get_db)],
    _ = Depends(allow_teacher)
):
    data = await get_report_data(class_id, start_date, end_date, report_type, db)
    return data

# --- 2. EXCEL ЭКСПОРТ ---
@router.get("/export")
async def export_report(
    class_id: int,
    report_type: str,
    start_date: date,
    end_date: date,
    db: Annotated[AsyncSession, Depends(get_db)],
    _ = Depends(allow_teacher)
):
    # 1. Получаем данные
    data = await get_report_data(class_id, start_date, end_date, report_type, db)
    class_info = await db.get(ClassGroup, class_id)
    class_name = class_info.name if class_info else "Unknown"

    # 2. Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    # Стили
    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Заголовок
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Отчет: {('Успеваемость' if report_type == 'grades' else 'Посещаемость')} | Класс: {class_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:D2')
    ws['A2'] = f"Период: {start_date} — {end_date}"
    ws['A2'].alignment = center_align

    # Шапка
    headers = ["№", "ФИО Ученика"]
    if report_type == "grades":
        headers.extend(["Средний балл", "Кол-во оценок"])
    else:
        headers.extend(["Пропуски (Н/Б)", "Опоздания"])

    ws.append([]) 
    ws.append(headers) 

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # Данные
    for idx, row in enumerate(data, 1):
        excel_row = [idx, row["full_name"]]
        if report_type == "grades":
            excel_row.append(row["value"])
            excel_row.append(row["count"])
        else:
            excel_row.append(row["absent"])
            excel_row.append(row["late"])
        ws.append(excel_row)
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=4 + idx, column=col_num).border = thin_border

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 3. 👇 ИСПРАВЛЕННАЯ ЛОГИКА ИМЕНИ ФАЙЛА 👇
    filename = f"Report_{class_name}_{report_type}_{start_date}.xlsx"
    encoded_filename = quote(filename)  # Кодируем русские буквы в %D0%90...
    
    return StreamingResponse(
        output, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={encoded_filename}"}
    )